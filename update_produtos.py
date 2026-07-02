#!/usr/bin/env python3
"""
update_produtos.py — AgroFito
──────────────────────────────
Exporta as 4 tabelas de Produtos Fitofarmacêuticos do SIFITO.
Com retry automático por tabela e tolerância a falhas parciais.
"""

import asyncio, json, sys, tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

TIMEOUT_NAV  = 180_000
TIMEOUT_ELEM = 120_000
TIMEOUT_DL   = 180_000
MAX_RETRIES  = 3
RETRY_DELAY  = 30
BASE_DIR     = Path(__file__).parent

COL_MAP = {
    0:  "designacao",
    1:  "autorizacao",
    2:  "numero",
    3:  "titular",
    4:  "tipo_util",
    5:  "substancia",
    6:  "teor",
    7:  "percentagem",
    8:  "formulacao",
    10: "classificacao",
    11: "frases",
    12: "baixo_risco",
    13: "cand_subs",
    14: "mpb",
    15: "data_autorizacao",
}

FUNC_TYPES = [
    'Fungicida','Inseticida','Herbicida','Acaricida','Moluscicida',
    'Nematodicida','Rodenticida','Fumigante','Regulador','Adjuvante',
    'Feromona','Bactericida','Algicida','Desinfestante','Repelente',
]

SOURCES = [
    dict(url="https://sifito.dgav.pt/divulgacao/produtos",
         estado="Autorizada", key="AUT",
         output="prod_autorizadas.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendapermitida",
         estado="Cancelada — Venda Permitida", key="CVP",
         output="prod_canceladas_venda_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendainterditautilizacaopermitida",
         estado="Cancelada — Venda Interdita / Util. Permitida", key="CVIP",
         output="prod_canceladas_venda_interdita_util_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendautilizacaointerdita",
         estado="Cancelada — Venda e Util. Interditas", key="CVUI",
         output="prod_canceladas_venda_util_interditas.json"),
]


def extract_funcao_tipo(s):
    up = (s or '').upper()
    for t in FUNC_TYPES:
        if t.upper() in up:
            return t
    return s.split()[0].slice(0, 25) if s.strip() else ''


async def download_xlsx(page, url: str) -> bytes:
    print(f"   🌐  {url}")
    await page.goto(url, timeout=TIMEOUT_NAV, wait_until="load")
    await page.locator(".k-grid").wait_for(timeout=TIMEOUT_ELEM)

    btn = page.locator(
        "button",
        has=page.locator("span.k-button-text", has_text="Exportar para Excel")
    ).first
    await btn.wait_for(timeout=TIMEOUT_ELEM)

    # Clica no botão para abrir eventual submenu
    await btn.click()

    # Tenta clicar no item do submenu se aparecer
    menu_item = page.locator("span.k-menu-link-text", has_text="Exportar para Excel").first
    try:
        await menu_item.wait_for(timeout=5_000)
        print("   📂  Submenu detectado — a clicar no item de download…")
        async with page.expect_download(timeout=TIMEOUT_DL) as dl_info:
            await menu_item.click()
    except Exception:
        print("   📂  Download directo…")
        async with page.expect_download(timeout=TIMEOUT_DL) as dl_info:
            await btn.click()

    download = await dl_info.value
    print(f"   📥  {download.suggested_filename}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        await download.save_as(tmp.name)
        data = Path(tmp.name).read_bytes()
        Path(tmp.name).unlink()
    return data


async def download_with_retry(browser, url: str) -> bytes:
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()
        try:
            if attempt > 1:
                print(f"   🔄  Tentativa {attempt}/{MAX_RETRIES}…")
            data = await download_xlsx(page, url)
            await context.close()
            return data
        except Exception as e:
            last_err = e
            print(f"   ⚠️  Tentativa {attempt} falhou: {e}")
            await context.close()
            if attempt < MAX_RETRIES:
                print(f"   ⏳  A aguardar {RETRY_DELAY}s antes de tentar novamente…")
                await asyncio.sleep(RETRY_DELAY)
    raise RuntimeError(f"Todas as tentativas falharam: {last_err}")


def xlsx_to_records(xlsx_bytes: bytes, estado: str) -> list[dict]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes); tmp_path = tmp.name
    wb = load_workbook(tmp_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    Path(tmp_path).unlink()

    records = []
    for row in rows[1:]:   # linha 0 = cabeçalho, dados de 1 em diante
        obj = {"estado": estado}
        for idx, key in COL_MAP.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                obj[key] = ""
            elif isinstance(val, datetime):
                obj[key] = val.strftime("%Y-%m-%d")
            else:
                v = str(val).strip()
                obj[key] = "" if v in ("-", "- -", "--") else v
        funcao_raw = str(row[9] or '').strip() if len(row) > 9 else ''
        parts = funcao_raw.split('|')
        obj['funcao_curta']     = parts[0].strip() if parts else ''
        obj['funcao_mecanismo'] = parts[1].strip() if len(parts) > 1 else ''
        up = obj['funcao_curta'].upper()
        obj['funcao_tipo'] = next((t for t in FUNC_TYPES if t.upper() in up), '')
        records.append(obj)
    return records


def save_json(records, output, date):
    path = BASE_DIR / output
    path.write_text(
        json.dumps({"date": date, "records": records},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    sz = path.stat().st_size / 1_048_576
    print(f"   💾  {output}  ({len(records):,} registos · {sz:.1f} MB)")


async def main():
    today   = datetime.now().strftime("%d/%m/%Y")
    failed  = []
    success = []
    print("=" * 55)
    print("  AgroFito — Produtos Fitofarmacêuticos (4 tabelas)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        for i, src in enumerate(SOURCES, 1):
            print(f"\n[{i}/4] {src['estado']}")
            try:
                xlsx = await download_with_retry(browser, src["url"])
                recs = xlsx_to_records(xlsx, src["estado"])
                save_json(recs, src["output"], today)
                success.append(src["key"])
            except Exception as e:
                print(f"   ❌  Falhou definitivamente: {e}", file=sys.stderr)
                failed.append(src["key"])
        await browser.close()
    print(f"\n{'='*55}")
    print(f"  Concluído: {len(success)}/4 tabelas actualizadas")
    if failed:
        print(f"  Falharam: {', '.join(failed)}")
    if len(failed) == len(SOURCES):
        print("\n❌  Nenhuma tabela foi actualizada.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
