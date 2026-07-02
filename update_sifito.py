#!/usr/bin/env python3
"""
update_sifito.py — AgroFito
────────────────────────────
Exporta as 4 tabelas de Condições de Utilização do SIFITO.

Retry automático: se o SIFITO estiver instável, tenta 3 vezes
por tabela antes de desistir. Continua para as restantes mesmo
que uma falhe.
"""

import asyncio, json, sys, tempfile, time
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

TIMEOUT_NAV  = 180_000   # 3 min para navegar
TIMEOUT_ELEM = 120_000   # 2 min para esperar elementos
TIMEOUT_DL   = 180_000   # 3 min para download
MAX_RETRIES  = 3         # tentativas por tabela
RETRY_DELAY  = 30        # segundos entre tentativas
BASE_DIR     = Path(__file__).parent

COL_MAP = {
    0:  "cultura",
    3:  "sit_particular",
    4:  "ambiente",
    5:  "inimigo",
    6:  "nome_cient",
    8:  "uso_menor",
    12: "produto",
    13: "autorizacao",
    14: "numero",
    15: "funcao",
    16: "substancia",
    17: "epoca",
    18: "tecnica",
    19: "num_max_intervalo",
    20: "concentracao",
    21: "vol_calda",
    22: "dose",
    23: "intervalo_seg",
    24: "restricoes",
    26: "validade",
    27: "limite_comerc",
    28: "limite_util",
}

SOURCES = [
    dict(url="https://sifito.dgav.pt/divulgacao/usos",
         estado="Autorizada", key="AUT",
         output="data_autorizadas.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/usoscanceladosvendapermitida",
         estado="Cancelada — Venda Permitida", key="CVP",
         output="data_canceladas_venda_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/usoscanceladosvendainterditautilizacaopermitida",
         estado="Cancelada — Venda Interdita / Util. Permitida", key="CVIP",
         output="data_canceladas_venda_interdita_util_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/usoscanceladosvendautilizacaointerdita",
         estado="Cancelada — Venda e Util. Interditas", key="CVUI",
         output="data_canceladas_venda_util_interditas.json"),
]


async def download_xlsx(page, url: str) -> bytes:
    """Navega para url e descarrega o Excel."""
    print(f"   🌐  {url}")
    await page.goto(url, timeout=TIMEOUT_NAV, wait_until="load")

    # Aguarda o contentor da grelha
    await page.locator(".k-grid").wait_for(timeout=TIMEOUT_ELEM)

    # Aguarda que os dados carreguem (pelo menos uma linha visível)
    print(f"   ⏳  A aguardar dados da tabela…")
    await page.locator(".k-grid-table tr").first.wait_for(timeout=TIMEOUT_ELEM)
    print(f"   ✅  Dados carregados.")

    # Clica no botão para abrir eventual submenu
    btn = page.locator(
        "button",
        has=page.locator("span.k-button-text", has_text="Exportar para Excel")
    ).first
    await btn.wait_for(timeout=TIMEOUT_ELEM)
    await btn.click()

    # Tenta clicar no item do submenu se aparecer
    menu_item = page.locator("span.k-menu-link-text", has_text="Exportar para Excel").first
    try:
        await menu_item.wait_for(timeout=5_000)
        print("   📂  Submenu detectado — a clicar no item de download…")
        async with page.expect_download(timeout=TIMEOUT_DL) as dl_info:
            await menu_item.click()
    except Exception:
        print("   📂  Download directo…")
        async with page.expect_download(timeout=TIMEOUT_DL) as dl_info:
            await btn.click()

    download = await dl_info.value
    print(f"   📥  {download.suggested_filename}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        await download.save_as(tmp.name)
        data = Path(tmp.name).read_bytes()
        Path(tmp.name).unlink()
    return data


async def download_with_retry(browser, url: str) -> bytes:
    """Tenta download até MAX_RETRIES vezes, criando nova página em cada tentativa."""
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()
        try:
            if attempt > 1:
                print(f"   🔄  Tentativa {attempt}/{MAX_RETRIES}…")
            data = await download_xlsx(page, url)
            await context.close()
            return data
        except Exception as e:
            last_err = e
            print(f"   ⚠️  Tentativa {attempt} falhou: {e}")
            await context.close()
            if attempt < MAX_RETRIES:
                print(f"   ⏳  A aguardar {RETRY_DELAY}s antes de tentar novamente…")
                await asyncio.sleep(RETRY_DELAY)
    raise RuntimeError(f"Todas as tentativas falharam: {last_err}")


def xlsx_to_records(xlsx_bytes: bytes, estado: str) -> list[dict]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes); tmp_path = tmp.name
    wb   = load_workbook(tmp_path, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    Path(tmp_path).unlink()

    records = []
    for row in rows[2:]:
        obj = {"estado": estado}
        for idx, key in COL_MAP.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                obj[key] = ""
            elif isinstance(val, datetime):
                obj[key] = val.strftime("%Y-%m-%d")
            else:
                v = str(val).strip()
                obj[key] = "" if v == "-" else v
        records.append(obj)
    return records


def save_json(records, output, date):
    path = BASE_DIR / output
    path.write_text(
        json.dumps({"date": date, "records": records},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    sz = path.stat().st_size / 1_048_576
    print(f"   💾  {output}  ({len(records):,} registos · {sz:.1f} MB)")


async def main():
    today   = datetime.now().strftime("%d/%m/%Y")
    failed  = []
    success = []

    print("=" * 55)
    print("  AgroFito — Condições de Utilização (4 tabelas)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        for i, src in enumerate(SOURCES, 1):
            print(f"\n[{i}/4] {src['estado']}")
            try:
                xlsx  = await download_with_retry(browser, src["url"])
                recs  = xlsx_to_records(xlsx, src["estado"])
                save_json(recs, src["output"], today)
                success.append(src["key"])
            except Exception as e:
                print(f"   ❌  Falhou definitivamente: {e}", file=sys.stderr)
                failed.append(src["key"])

        await browser.close()

    print(f"\n{'='*55}")
    print(f"  Concluído: {len(success)}/4 tabelas actualizadas")
    if failed:
        print(f"  Falharam: {', '.join(failed)}")
    print(f"{'='*55}")

    # Falha o workflow só se TODAS as tabelas falharam
    if len(failed) == len(SOURCES):
        print("\n❌  Nenhuma tabela foi actualizada.", file=sys.stderr)
        sys.exit(1)

    if failed:
        print(f"\n⚠️  {len(failed)} tabela(s) não actualizadas — as restantes foram guardadas.")


if __name__ == "__main__":
    asyncio.run(main())
